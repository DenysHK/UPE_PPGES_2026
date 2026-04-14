
"""
Script para ler a planilha RSB_Results.xlsx, acessar a aba
"101325 Pa EXP (49)", reproduzir exatamente a escolha dos seis
casos R1..R6 e criar uma nova aba chamada "SEIS ESCOLHIDOS".

BIBLIOTECAS UTILIZADAS
----------------------
- pandas: para organizar os casos em DataFrame e aplicar os critérios
  de seleção de forma explícita e reprodutível;
- openpyxl: para ler a estrutura da planilha original (que está em formato
  "largo", com uma simulação por coluna) e para escrever a nova aba.

CRITÉRIO FORMALIZADO
--------------------
1) Universo de candidatos:
   - usar somente a aba "101325 Pa EXP (49)";
   - usar somente as simulações a 1 atm;
   - restringir a escolha às composições intermediárias de CH4:
     0,50; 0,525; 0,55; 0,575; 0,60.

2) Faixa-alvo definida a partir da SFT:
   - limite inferior = 1,80
   - centro da faixa = 1,90
   - limite superior = 2,00

3) Regras para reproduzir exatamente R1..R6:
   - R1: para CH4 = 0,50, escolher o caso com H2/CO > 2,00
         mais próximo de 2,00  -> "âncora superior";
   - R2: para CH4 = 0,50, escolher o caso com 1,80 <= H2/CO <= 2,00
         mais próximo de 1,90  -> representante central;
   - R3: para CH4 = 0,525, escolher o caso mais próximo de 1,90;
   - R4: para CH4 = 0,55, escolher o caso mais próximo de 1,90;
   - R5: para CH4 = 0,575, escolher o caso com H2/CO >= 1,80
         mais próximo de 1,80  -> "âncora inferior";
   - R6: para CH4 = 0,60, escolher o caso mais próximo de 1,90.

OBSERVAÇÃO IMPORTANTE
---------------------
Esse critério NÃO significa "pegar sempre o mais próximo de 1,90".
Ele foi construído para reproduzir exatamente o conjunto final:
- uma âncora superior (R1),
- pontos centrais (R2, R3, R4 e R6),
- uma âncora inferior (R5).

Assim, por exemplo, para CH4 = 0,575, o caso 1073 K com H2/CO = 1,92061
é mais próximo do centro 1,90, mas não foi escolhido porque o papel de R5
é representar a borda inferior da faixa promissora, e não mais um ponto central.
"""

from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ======================================================================
# 1) DEFINIÇÃO DE ARQUIVOS E PARÂMETROS
# ======================================================================

# Arquivo de entrada.
INPUT_FILE = Path("/mnt/data/RSB_Results.xlsx")

# Arquivo de saída:
# para preservar o original, o script grava uma cópia com a nova aba.
OUTPUT_FILE = Path("/mnt/data/RSB_Results_com_SEIS_ESCOLHIDOS.xlsx")

# Aba da qual os dados serão lidos.
SOURCE_SHEET = "101325 Pa EXP (49)"

# Nome da nova aba que será criada.
TARGET_SHEET = "SEIS ESCOLHIDOS"

# Valores da faixa-alvo vindos da SFT.
H2CO_LOW = 1.80
H2CO_CENTER = 1.90
H2CO_UP = 2.00

# Composições intermediárias consideradas na seleção.
INTERMEDIATE_CH4 = [0.50, 0.525, 0.55, 0.575, 0.60]

# ======================================================================
# 2) FUNÇÕES AUXILIARES
# ======================================================================

def parse_temperature(value):
    """
    Converte textos como '973 K' em inteiro 973.
    Se o valor já for numérico, retorna inteiro.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"(\d+)", str(value))
    return int(match.group(1)) if match else None


def safe_float(value):
    """
    Converte o valor para float quando possível.
    Retorna None se a célula estiver vazia.
    """
    if value is None:
        return None
    return float(value)


def pick_closest(df, target, label):
    """
    Escolhe o caso cujo H2/CO está mais próximo de um alvo numérico.
    Em caso de empate, escolhe a menor temperatura para manter
    o resultado determinístico.
    """
    if df.empty:
        raise ValueError(f"Nenhum candidato disponível para {label}.")
    tmp = df.copy()
    tmp["distance"] = (tmp["H2_CO"] - target).abs()
    tmp = tmp.sort_values(["distance", "Temperature_K"], ascending=[True, True])
    chosen = tmp.iloc[0].copy()
    chosen["criterion_target"] = target
    chosen["criterion_distance"] = abs(chosen["H2_CO"] - target)
    return chosen


# ======================================================================
# 3) LEITURA DA PLANILHA E CONVERSÃO DA ABA LARGA EM TABELA
# ======================================================================

# data_only=True:
# lê os valores calculados das fórmulas (por exemplo, H2/CO),
# o que é essencial para aplicar os critérios de seleção.
wb_values = load_workbook(INPUT_FILE, data_only=True)
ws_values = wb_values[SOURCE_SHEET]

# Workbook separado para edição/escrita.
wb_edit = load_workbook(INPUT_FILE)
ws_edit_source = wb_edit[SOURCE_SHEET]

# A aba está estruturada com uma simulação por coluna.
# Algumas colunas são vazias (separadores entre blocos de temperatura),
# então vamos varrer todas as colunas e coletar apenas aquelas que
# realmente representam simulações válidas.
records = []
last_temperature = None

for col in range(3, ws_values.max_column + 1):
    # Temperatura da linha 2. Em algumas colunas pode vir vazia
    # por causa da formatação da planilha; nesse caso usamos a última
    # temperatura válida ("forward fill").
    temp_raw = ws_values.cell(2, col).value
    temp = parse_temperature(temp_raw)
    if temp is not None:
        last_temperature = temp
    else:
        temp = last_temperature

    ch4 = ws_values.cell(8, col).value
    co2 = ws_values.cell(9, col).value
    h2co = ws_values.cell(32, col).value

    # Se não houver composição ou H2/CO, a coluna não representa um caso válido.
    if ch4 is None or h2co is None:
        continue

    record = {
        "Column": get_column_letter(col),
        "Temperature_K": temp,
        "Pressure_Pa": safe_float(ws_values.cell(3, col).value),
        "Biogas_Flow_mol_s": safe_float(ws_values.cell(4, col).value),
        "CH4_frac": safe_float(ch4),
        "CO2_frac": safe_float(co2),
        "CH4_mol_s": safe_float(ws_values.cell(12, col).value),
        "CO2_mol_s": safe_float(ws_values.cell(13, col).value),
        "CO_mol_s": safe_float(ws_values.cell(14, col).value),
        "H2O_mol_s": safe_float(ws_values.cell(15, col).value),
        "C_mol_s": safe_float(ws_values.cell(16, col).value),
        "H2_mol_s": safe_float(ws_values.cell(17, col).value),
        "OUT_mol_s": safe_float(ws_values.cell(18, col).value),
        "DRM_mol_s": safe_float(ws_values.cell(20, col).value),
        "RWGS_mol_s": safe_float(ws_values.cell(23, col).value),
        "DM_mol_s": safe_float(ws_values.cell(26, col).value),
        "Conversion_CH4_pct": safe_float(ws_values.cell(29, col).value),
        "Conversion_CO2_pct": safe_float(ws_values.cell(30, col).value),
        "Selec_C": safe_float(ws_values.cell(31, col).value),
        "H2_CO": safe_float(h2co),
    }
    records.append(record)

df_all = pd.DataFrame(records)

# ======================================================================
# 4) FILTRO INICIAL: SOMENTE COMPOSIÇÕES INTERMEDIÁRIAS
# ======================================================================

df = df_all[df_all["CH4_frac"].isin(INTERMEDIATE_CH4)].copy()

# ======================================================================
# 5) APLICAÇÃO DOS CRITÉRIOS EXATOS PARA R1..R6
# ======================================================================

# R1 = âncora superior:
# composição 0,50 e H2/CO > 2,00; escolher o mais próximo de 2,00.
R1 = pick_closest(
    df[(df["CH4_frac"] == 0.50) & (df["H2_CO"] > H2CO_UP)],
    H2CO_UP,
    "R1"
)
R1["Case"] = "R1"
R1["Rule"] = "CH4=0,50; H2/CO>2,00; menor |H2/CO-2,00|"

# R2 = representante central para CH4 = 0,50:
# dentro da faixa [1,80; 2,00], escolher o mais próximo de 1,90.
R2 = pick_closest(
    df[(df["CH4_frac"] == 0.50) & (df["H2_CO"] >= H2CO_LOW) & (df["H2_CO"] <= H2CO_UP)],
    H2CO_CENTER,
    "R2"
)
R2["Case"] = "R2"
R2["Rule"] = "CH4=0,50; 1,80<=H2/CO<=2,00; menor |H2/CO-1,90|"

# R3 = representante central para CH4 = 0,525.
R3 = pick_closest(
    df[df["CH4_frac"] == 0.525],
    H2CO_CENTER,
    "R3"
)
R3["Case"] = "R3"
R3["Rule"] = "CH4=0,525; menor |H2/CO-1,90|"

# R4 = representante central para CH4 = 0,55.
R4 = pick_closest(
    df[df["CH4_frac"] == 0.55],
    H2CO_CENTER,
    "R4"
)
R4["Case"] = "R4"
R4["Rule"] = "CH4=0,55; menor |H2/CO-1,90|"

# R5 = âncora inferior:
# para CH4 = 0,575, escolher o caso >= 1,80 mais próximo de 1,80.
# Isso força a representação do limite inferior da faixa útil,
# em vez de gerar mais um ponto central redundante.
R5 = pick_closest(
    df[(df["CH4_frac"] == 0.575) & (df["H2_CO"] >= H2CO_LOW)],
    H2CO_LOW,
    "R5"
)
R5["Case"] = "R5"
R5["Rule"] = "CH4=0,575; H2/CO>=1,80; menor |H2/CO-1,80|"

# R6 = representante central para CH4 = 0,60.
R6 = pick_closest(
    df[df["CH4_frac"] == 0.60],
    H2CO_CENTER,
    "R6"
)
R6["Case"] = "R6"
R6["Rule"] = "CH4=0,60; menor |H2/CO-1,90|"

selected = pd.DataFrame([R1, R2, R3, R4, R5, R6])

# Reordena as colunas para a tabela final.
selected = selected[
    [
        "Case", "Rule", "Column", "Temperature_K", "Pressure_Pa",
        "Biogas_Flow_mol_s", "CH4_frac", "CO2_frac", "H2_CO",
        "criterion_target", "criterion_distance",
        "CO_mol_s", "H2_mol_s", "H2O_mol_s", "C_mol_s",
        "Conversion_CH4_pct", "Conversion_CO2_pct", "Selec_C",
        "DRM_mol_s", "RWGS_mol_s", "DM_mol_s", "OUT_mol_s"
    ]
].copy()

# ======================================================================
# 6) CRIAÇÃO / RECRIAÇÃO DA ABA "SEIS ESCOLHIDOS"
# ======================================================================

if TARGET_SHEET in wb_edit.sheetnames:
    del wb_edit[TARGET_SHEET]

ws_new = wb_edit.create_sheet(TARGET_SHEET)

# Estilos básicos.
fill_title = PatternFill("solid", fgColor="1F4E78")
fill_header = PatternFill("solid", fgColor="D9EAF7")
fill_section = PatternFill("solid", fgColor="E2F0D9")
font_white_bold = Font(color="FFFFFF", bold=True)
font_bold = Font(bold=True)
thin = Side(style="thin", color="A6A6A6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ----------------------------------------------------------------------
# Bloco de título
# ----------------------------------------------------------------------
ws_new["A1"] = "SEIS ESCOLHIDOS PARA ACOPLAMENTO RSB–SFT"
ws_new["A1"].font = font_white_bold
ws_new["A1"].fill = fill_title
ws_new["A1"].alignment = Alignment(horizontal="center")
ws_new.merge_cells("A1:L1")

# ----------------------------------------------------------------------
# Bloco explicando os critérios, de forma legível e auditável
# ----------------------------------------------------------------------
criteria_lines = [
    "Origem dos dados: aba '101325 Pa EXP (49)' de RSB_Results.xlsx.",
    "Filtro inicial: somente composições intermediárias CH4 = 0,50; 0,525; 0,55; 0,575; 0,60.",
    "Faixa-alvo vinda da SFT: limite inferior = 1,80; centro = 1,90; limite superior = 2,00.",
    "R1: CH4=0,50 e H2/CO>2,00; escolher o mais próximo de 2,00 (âncora superior).",
    "R2: CH4=0,50 e 1,80<=H2/CO<=2,00; escolher o mais próximo de 1,90.",
    "R3: CH4=0,525; escolher o mais próximo de 1,90.",
    "R4: CH4=0,55; escolher o mais próximo de 1,90.",
    "R5: CH4=0,575 e H2/CO>=1,80; escolher o mais próximo de 1,80 (âncora inferior).",
    "R6: CH4=0,60; escolher o mais próximo de 1,90.",
    "Justificativa do critério: reproduzir exatamente um conjunto com borda superior, região central e borda inferior da faixa promissora."
]

start_row = 3
for i, line in enumerate(criteria_lines, start=start_row):
    ws_new.cell(i, 1).value = line
    if i == start_row:
        ws_new.cell(i, 1).font = font_bold
        ws_new.cell(i, 1).fill = fill_section
    ws_new.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)
    ws_new.cell(i, 1).alignment = Alignment(wrap_text=True)

# ----------------------------------------------------------------------
# Tabela com os seis escolhidos
# ----------------------------------------------------------------------
table_start = start_row + len(criteria_lines) + 2

headers = [
    "Caso", "Regra aplicada", "Coluna origem", "T (K)", "P (Pa)",
    "Biogás (mol/s)", "CH4", "CO2", "H2/CO",
    "Alvo", "|H2/CO - alvo|",
    "CO (mol/s)", "H2 (mol/s)", "H2O (mol/s)", "C (mol/s)",
    "Conv. CH4 (%)", "Conv. CO2 (%)", "Selec(C)",
    "DRM (mol/s)", "RWGS (mol/s)", "DM (mol/s)", "OUT total (mol/s)"
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws_new.cell(table_start, col_idx)
    cell.value = header
    cell.font = font_bold
    cell.fill = fill_header
    cell.border = border
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for row_idx, (_, row) in enumerate(selected.iterrows(), start=table_start + 1):
    values = [
        row["Case"], row["Rule"], row["Column"], row["Temperature_K"], row["Pressure_Pa"],
        row["Biogas_Flow_mol_s"], row["CH4_frac"], row["CO2_frac"], row["H2_CO"],
        row["criterion_target"], row["criterion_distance"],
        row["CO_mol_s"], row["H2_mol_s"], row["H2O_mol_s"], row["C_mol_s"],
        row["Conversion_CH4_pct"], row["Conversion_CO2_pct"], row["Selec_C"],
        row["DRM_mol_s"], row["RWGS_mol_s"], row["DM_mol_s"], row["OUT_mol_s"]
    ]
    for col_idx, value in enumerate(values, start=1):
        cell = ws_new.cell(row_idx, col_idx)
        cell.value = value
        cell.border = border

# ----------------------------------------------------------------------
# Formatação numérica e visual da tabela
# ----------------------------------------------------------------------
# Colunas de números com 5 casas para facilitar auditoria.
for col in range(6, 23):
    for row in range(table_start + 1, table_start + 1 + len(selected)):
        ws_new.cell(row, col).number_format = "0.00000"

# Colunas de temperatura e pressão.
for row in range(table_start + 1, table_start + 1 + len(selected)):
    ws_new.cell(row, 4).number_format = "0"
    ws_new.cell(row, 5).number_format = "0"

# Ajuste de larguras de coluna.
widths = {
    "A": 8, "B": 52, "C": 14, "D": 10, "E": 14,
    "F": 16, "G": 10, "H": 10, "I": 12, "J": 10, "K": 14,
    "L": 14, "M": 14, "N": 14, "O": 14, "P": 14, "Q": 14,
    "R": 12, "S": 14, "T": 14, "U": 14, "V": 16
}
for col_letter, width in widths.items():
    ws_new.column_dimensions[col_letter].width = width

# Congela painel abaixo do cabeçalho da tabela.
ws_new.freeze_panes = ws_new[f"A{table_start + 1}"]

# Observação final.
note_row = table_start + 1 + len(selected) + 2
ws_new.cell(note_row, 1).value = (
    "Observação: o critério acima foi definido para reproduzir exatamente R1..R6, "
    "incluindo uma âncora superior (R1), casos centrais (R2, R3, R4, R6) "
    "e uma âncora inferior (R5)."
)
ws_new.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=12)
ws_new.cell(note_row, 1).alignment = Alignment(wrap_text=True)
ws_new.cell(note_row, 1).fill = fill_section

# Salva a nova planilha.
wb_edit.save(OUTPUT_FILE)

print(f"Arquivo salvo em: {OUTPUT_FILE}")
